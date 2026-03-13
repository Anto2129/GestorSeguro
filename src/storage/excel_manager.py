"""Gestión de archivos Excel"""
from pathlib import Path
from typing import List, Set
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from config.settings import AppConfig
from models.usuario import Usuario
from ui.colors import Colores


class GestorExcel:
    """Gestiona archivos Excel de usuarios"""
    
    def __init__(self, config: AppConfig):
        self.config = config
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """Crea archivo si no existe"""
        if not self.config.archivo_log.exists():
            self._create_formatted_file()
        else:
            print(f"{Colores.VERDE}✓ Archivo encontrado{Colores.RESET}\n")
    
    def _create_formatted_file(self):
        """Crea archivo Excel formateado"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        headers = [
            "Nombre de usuario",
            "Contraseña Hash",
            "Fecha de creacion",
            "uso",
            "Estado",
            "Dispositivo",
            "Fecha Modificación",
            "Notas"
        ]
        
        # Estilos
        font_header = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        fill_header = PatternFill(
            start_color="2180a0",
            end_color="2180a0",
            fill_type="solid"
        )
        alignment_header = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment_header
            cell.border = border
        
        # Ancho de columnas
        widths = [20, 65, 20, 25, 18, 30, 20, 30]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        ws.row_dimensions[1].height = 25
        
        # Data validation para Estado
        dv = DataValidation(
            type="list",
            formula1='"Activo,Inactivo"',
            allow_blank=False
        )
        dv.error = 'Selecciona "Activo" o "Inactivo"'
        ws.add_data_validation(dv)
        dv.add('E2:E1000')
        
        wb.save(self.config.archivo_log)
        print(f"{Colores.VERDE}✓ Archivo creado{Colores.RESET}\n")
    
    def get_all_users(self) -> List[Usuario]:
        """Obtiene todos los usuarios"""
        usuarios = []
        
        try:
            wb = load_workbook(self.config.archivo_log)
            ws = wb.active
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row and row:
                    usuarios.append(Usuario.from_excel_row(row))
        
        except Exception as e:
            print(f"{Colores.ROJO}Error: {e}{Colores.RESET}")
        
        return usuarios
    
    def get_existing_usernames(self) -> Set[str]:
        """Obtiene nombres de usuarios existentes"""
        return set(u.nombre for u in self.get_all_users())
    
    def get_existing_password_hashes(self) -> Set[str]:
        """Obtiene hashes de contraseñas existentes"""
        return set(u.contraseña_hash for u in self.get_all_users())
    
    def add_user(self, usuario: Usuario) -> bool:
        """Añade usuario a Excel"""
        try:
            wb = load_workbook(self.config.archivo_log)
            ws = wb.active
            
            next_row = ws.max_row + 1
            datos = usuario.to_excel_row()
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col_num, dato in enumerate(datos, 1):
                cell = ws.cell(row=next_row, column=col_num)
                cell.value = dato
                cell.border = border
            
            ws.row_dimensions[next_row].height = 20
            wb.save(self.config.archivo_log)
            
            return True
        
        except Exception as e:
            print(f"{Colores.ROJO}Error: {e}{Colores.RESET}")
            return False
    
    def update_user_status(self, username: str, new_status: str) -> bool:
        """Actualiza estado de usuario"""
        try:
            wb = load_workbook(self.config.archivo_log)
            ws = wb.active
            
            from utils.date_utils import FechaUtils
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == username:
                    row[4].value = new_status
                    row[6].value = FechaUtils.get_current_datetime_formatted()
                    wb.save(self.config.archivo_log)
                    return True
            
            return False
        
        except Exception as e:
            print(f"{Colores.ROJO}Error: {e}{Colores.RESET}")
            return False
    
    def delete_user(self, username: str) -> bool:
        """Borra usuario de Excel"""
        try:
            wb = load_workbook(self.config.archivo_log)
            ws = wb.active
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == username:
                    ws.delete_rows(row_num, 1)
                    wb.save(self.config.archivo_log)
                    return True
            
            return False
        
        except Exception as e:
            print(f"{Colores.ROJO}Error: {e}{Colores.RESET}")
            return False